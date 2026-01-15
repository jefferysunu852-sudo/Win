import re
from typing import List, Optional, Any, Dict, Tuple
from openpyxl.worksheet.worksheet import Worksheet

class ParsedRow:
    """Structure to hold parsed row data with section context."""
    def __init__(self, row_idx: int, section: str, raw_key: str, key: str, values: Dict[str, Any]):
        self.row_idx = row_idx
        self.section = section
        self.raw_key = raw_key
        self.key = key          # Normalized Material Name
        self.values = values    # Dict of column_alias -> value

    def composite_key(self) -> Tuple[str, str]:
        return (self.section, self.key)

    def __repr__(self):
        return f"<Row {self.row_idx} [{self.section}] {self.key}: {self.values}>"

def normalize_string(s: str) -> str:
    """Standardizes strings for comparison: lower, strip, remove punctuation."""
    if not s:
        return ""
    # Casefold and strip
    s = str(s).lower()
    # Keep alphanumeric and spaces
    s = re.sub(r'[^a-z0-9\s]', '', s)
    # Collapse whitespace
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def parse_number(val: Any) -> Optional[float]:
    """
    Parses numbers from potential strings.
    - 1,5 -> 1.5
    - 1 234,5 -> 1234.5
    - None/Blank -> None
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    
    s = str(val).strip()
    if not s:
        return None
        
    try:
        # Handle spaces as thousand separators
        s = s.replace(" ", "")
        # Handle comma as decimal
        s = s.replace(",", ".")
        return float(s)
    except ValueError:
        return None

class SectionParser:
    """
    Parses a sheet considering 'Work Sections'.
    
    A row is a 'Section Header' if:
    1. Key Column (C) has text.
    2. Data Columns (specified range) are EMPTY or Non-Numeric.
    """
    def __init__(self, sheet: Worksheet, key_col_idx: int = 3, start_row: int = 14):
        self.sheet = sheet
        self.key_col_idx = key_col_idx
        self.start_row = start_row

    def parse(self, data_cols: List[int], value_mapping: Dict[str, int], detect_sections: bool = True) -> List[ParsedRow]:
        """
        data_cols: List of column indices to check for emptiness (to detect Section Header).
        value_mapping: Dict of { 'alias': col_idx } to extract data from.
        detect_sections: If False, treats every row as a data row.
        """
        results = []
        current_section = "__NO_SECTION__"
        
        # Determine strict max row to avoid iterating millions of empty rows
        max_r = self.sheet.max_row
        
        for r in range(self.start_row, max_r + 1):
            # Read Key
            key_cell = self.sheet.cell(row=r, column=self.key_col_idx)
            raw_key = str(key_cell.value) if key_cell.value else ""
            key_norm = normalize_string(raw_key)
            
            if not key_norm:
                continue
            
            # Check if Section Header
            is_section_header = False
            if detect_sections:
                is_header_candidate = True
                for col in data_cols:
                    val = self.sheet.cell(row=r, column=col).value
                    # If we find a number, it's a data row
                    num = parse_number(val)
                    if num is not None:
                        is_header_candidate = False
                        break
                if is_header_candidate:
                    is_section_header = True

            if is_section_header:
                current_section = re.sub(r'\s+', ' ', raw_key).strip()
                continue
            
            # It is a Data Row
            row_values = {}
            for alias, col_idx in value_mapping.items():
                val = self.sheet.cell(row=r, column=col_idx).value
                p_val = parse_number(val)
                row_values[alias] = p_val
            
            results.append(ParsedRow(r, current_section, raw_key, key_norm, row_values))
            
        return results
