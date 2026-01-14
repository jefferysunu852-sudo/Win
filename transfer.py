from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from excel_layout import (
    WeekBlock,
    find_data_start_row,
    normalize_text,
    parse_number,
)


@dataclass
class TransferSettings:
    write_all_duplicates: bool = False
    overwrite_formulas: bool = False


@dataclass
class TransferResult:
    diff_rows: List[Dict[str, object]]
    logs: List[str]
    written_cells: int
    matched_keys: int
    missing_target_keys: int
    duplicate_source_keys: int
    duplicate_target_keys: int
    skipped_formula_cells: int


def _build_key_map(ws: Worksheet, start_row: int) -> Dict[str, List[int]]:
    key_map: Dict[str, List[int]] = {}
    for row in range(start_row, ws.max_row + 1):
        key_value = ws.cell(row=row, column=3).value
        key = normalize_text(key_value)
        if not key:
            continue
        key_map.setdefault(key, []).append(row)
    return key_map


def _aggregate_source(ws: Worksheet, block: WeekBlock, start_row: int) -> Tuple[Dict[str, Dict[str, Optional[float]]], int]:
    data: Dict[str, Dict[str, Optional[float]]] = {}
    key_map = _build_key_map(ws, start_row)
    duplicate_keys = sum(1 for rows in key_map.values() if len(rows) > 1)

    for key, rows in key_map.items():
        totals = {"planned": None, "actual": None, "timesheet": None}
        for row in rows:
            planned = parse_number(ws.cell(row=row, column=block.start_col).value)
            actual = parse_number(ws.cell(row=row, column=block.start_col + 2).value)
            timesheet = parse_number(ws.cell(row=row, column=block.start_col + 4).value)
            if planned is not None:
                totals["planned"] = (totals["planned"] or 0.0) + planned
            if actual is not None:
                totals["actual"] = (totals["actual"] or 0.0) + actual
            if timesheet is not None:
                totals["timesheet"] = (totals["timesheet"] or 0.0) + timesheet
        data[key] = totals
    return data, duplicate_keys


def _is_formula_cell(cell) -> bool:
    return cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))


def analyze_transfer(
    source_ws: Worksheet,
    target_ws: Worksheet,
    source_block: WeekBlock,
    target_block: WeekBlock,
    settings: TransferSettings,
) -> TransferResult:
    source_start = find_data_start_row(source_ws)
    target_start = find_data_start_row(target_ws)
    source_data, duplicate_source = _aggregate_source(source_ws, source_block, source_start)
    target_map = _build_key_map(target_ws, target_start)

    diff_rows: List[Dict[str, object]] = []
    logs: List[str] = []
    skipped_formula_cells = 0
    missing_target = 0
    duplicate_target = sum(1 for rows in target_map.values() if len(rows) > 1)

    for key, totals in source_data.items():
        target_rows = target_map.get(key)
        if not target_rows:
            missing_target += 1
            diff_rows.append(
                {
                    "key": key,
                    "src_planned_qty": totals["planned"],
                    "src_actual_qty": totals["actual"],
                    "src_timesheet": totals["timesheet"],
                    "tgt_planned_qty": None,
                    "tgt_actual_qty": None,
                    "tgt_timesheet": None,
                    "action": "skip",
                    "reason": "missing target key",
                }
            )
            continue

        rows_to_show = target_rows if settings.write_all_duplicates else target_rows[:1]
        for row in rows_to_show:
            planned_cell = target_ws.cell(row=row, column=target_block.start_col)
            actual_cell = target_ws.cell(row=row, column=target_block.start_col + 2)
            timesheet_cell = target_ws.cell(row=row, column=target_block.start_col + 4)

            planned_value = planned_cell.value
            actual_value = actual_cell.value
            timesheet_value = timesheet_cell.value

            reasons = []
            writable_cells = 0
            if totals["planned"] is not None:
                if settings.overwrite_formulas or not _is_formula_cell(planned_cell):
                    writable_cells += 1
                elif not settings.overwrite_formulas:
                    reasons.append("planned has formula")
                    skipped_formula_cells += 1
            if totals["actual"] is not None:
                if settings.overwrite_formulas or not _is_formula_cell(actual_cell):
                    writable_cells += 1
                elif not settings.overwrite_formulas:
                    reasons.append("actual has formula")
                    skipped_formula_cells += 1
            if totals["timesheet"] is not None:
                if settings.overwrite_formulas or not _is_formula_cell(timesheet_cell):
                    writable_cells += 1
                elif not settings.overwrite_formulas:
                    reasons.append("timesheet has formula")
                    skipped_formula_cells += 1

            has_values = any(value is not None for value in totals.values())
            action = "write" if has_values and writable_cells else "skip"

            diff_rows.append(
                {
                    "key": key,
                    "src_planned_qty": totals["planned"],
                    "src_actual_qty": totals["actual"],
                    "src_timesheet": totals["timesheet"],
                    "tgt_planned_qty": planned_value,
                    "tgt_actual_qty": actual_value,
                    "tgt_timesheet": timesheet_value,
                    "action": action,
                    "reason": ", ".join(reasons) if reasons else "",
                }
            )

    return TransferResult(
        diff_rows=diff_rows,
        logs=logs,
        written_cells=0,
        matched_keys=len(source_data) - missing_target,
        missing_target_keys=missing_target,
        duplicate_source_keys=duplicate_source,
        duplicate_target_keys=duplicate_target,
        skipped_formula_cells=skipped_formula_cells,
    )


def transfer_week(
    source_ws: Worksheet,
    target_ws: Worksheet,
    source_block: WeekBlock,
    target_block: WeekBlock,
    settings: TransferSettings,
) -> TransferResult:
    source_start = find_data_start_row(source_ws)
    target_start = find_data_start_row(target_ws)
    source_data, duplicate_source = _aggregate_source(source_ws, source_block, source_start)
    target_map = _build_key_map(target_ws, target_start)

    diff_rows: List[Dict[str, object]] = []
    logs: List[str] = []
    written_cells = 0
    missing_target = 0
    duplicate_target = sum(1 for rows in target_map.values() if len(rows) > 1)
    skipped_formula_cells = 0

    for key, totals in source_data.items():
        target_rows = target_map.get(key)
        if not target_rows:
            missing_target += 1
            logs.append(f"Missing target key: {key}")
            diff_rows.append(
                {
                    "key": key,
                    "src_planned_qty": totals["planned"],
                    "src_actual_qty": totals["actual"],
                    "src_timesheet": totals["timesheet"],
                    "tgt_planned_qty": None,
                    "tgt_actual_qty": None,
                    "tgt_timesheet": None,
                    "action": "skip",
                    "reason": "missing target key",
                }
            )
            continue

        rows_to_write = target_rows if settings.write_all_duplicates else target_rows[:1]
        if len(target_rows) > 1 and not settings.write_all_duplicates:
            logs.append(f"Duplicate target key '{key}' - writing to first match")
        elif len(target_rows) > 1:
            logs.append(f"Duplicate target key '{key}' - writing to all matches")

        for row in rows_to_write:
            planned_cell = target_ws.cell(row=row, column=target_block.start_col)
            actual_cell = target_ws.cell(row=row, column=target_block.start_col + 2)
            timesheet_cell = target_ws.cell(row=row, column=target_block.start_col + 4)

            reasons = []
            writable_cells = 0
            if totals["planned"] is not None:
                if settings.overwrite_formulas or not _is_formula_cell(planned_cell):
                    writable_cells += 1
                elif not settings.overwrite_formulas:
                    reasons.append("planned has formula")
                    skipped_formula_cells += 1
            if totals["actual"] is not None:
                if settings.overwrite_formulas or not _is_formula_cell(actual_cell):
                    writable_cells += 1
                elif not settings.overwrite_formulas:
                    reasons.append("actual has formula")
                    skipped_formula_cells += 1
            if totals["timesheet"] is not None:
                if settings.overwrite_formulas or not _is_formula_cell(timesheet_cell):
                    writable_cells += 1
                elif not settings.overwrite_formulas:
                    reasons.append("timesheet has formula")
                    skipped_formula_cells += 1

            has_values = any(value is not None for value in totals.values())
            action = "write" if has_values and writable_cells else "skip"

            if action == "write":
                if totals["planned"] is not None and (settings.overwrite_formulas or not _is_formula_cell(planned_cell)):
                    planned_cell.value = float(totals["planned"])
                    written_cells += 1
                if totals["actual"] is not None and (settings.overwrite_formulas or not _is_formula_cell(actual_cell)):
                    actual_cell.value = float(totals["actual"])
                    written_cells += 1
                if totals["timesheet"] is not None and (settings.overwrite_formulas or not _is_formula_cell(timesheet_cell)):
                    timesheet_cell.value = float(totals["timesheet"])
                    written_cells += 1

            diff_rows.append(
                {
                    "key": key,
                    "src_planned_qty": totals["planned"],
                    "src_actual_qty": totals["actual"],
                    "src_timesheet": totals["timesheet"],
                    "tgt_planned_qty": planned_cell.value,
                    "tgt_actual_qty": actual_cell.value,
                    "tgt_timesheet": timesheet_cell.value,
                    "action": action,
                    "reason": ", ".join(reasons) if reasons else "",
                }
            )

    return TransferResult(
        diff_rows=diff_rows,
        logs=logs,
        written_cells=written_cells,
        matched_keys=len(source_data) - missing_target,
        missing_target_keys=missing_target,
        duplicate_source_keys=duplicate_source,
        duplicate_target_keys=duplicate_target,
        skipped_formula_cells=skipped_formula_cells,
    )
