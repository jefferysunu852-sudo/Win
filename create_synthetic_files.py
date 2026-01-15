from openpyxl import Workbook
from openpyxl.styles import Font

def create_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Year
    ws.merge_cells("A8:M8")
    ws["A8"] = "2024"
    
    # Key Column (C)
    ws["C13"] = "Material Name"
    
    # Week 1 (Starts at H - Col 8)
    # H, I, J, K, L
    ws.merge_cells("H10:L10")
    ws["H10"] = "Week 1 - Jan 1"
    
    # Subheaders
    headers = ["q-ty (plan)", "M/hour (plan)", "q-ty (actual)", "Man/hour (actual)", "Timesheet"]
    for i, h in enumerate(headers):
        ws.cell(row=13, column=8+i, value=h)
        
    # Data
    # Section 1
    ws["C14"] = "Section Alpha" # Header
    # H14:L14 should be empty for a section header
    
    ws["C15"] = "Concrete"
    ws["C15"].font = Font(bold=False)
    # Give Concrete data so it isn't seen as a Section Header
    ws["H15"] = 200
    
    ws["C16"] = "Steel"
    
    # Values for Steel
    # Plan (Col H - 8)
    ws["H16"] = 100 
    # Actual (Col J - 10)
    ws["J16"] = 50
    # Timesheet (Col L - 12)
    ws["L16"] = 10
    
    wb.save("synthetic_report.xlsx")

def create_cost_control():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Control"
    
    # Week Block 1: Same position (H-L)
    ws.merge_cells("H10:L10")
    ws["H10"] = "Week 1 - Jan 1"
    headers = ["q-ty", "mh", "q-ty", "mh", "timesheet"]
    for i, h in enumerate(headers):
        ws.cell(row=13, column=8+i, value=h)
    
    # Action 2 Source Structure
    # "From beginning... Q-ty" in Source usually is further right or somewhere else.
    # Let's say it's at M10:N11 (Col 13/14).
    # Prompt said "H10:I11 merged contains text...". 
    # The prompt might imply Week blocks are somewhere else or the file structure is different for Action 2 source?
    # "Action 2 Source: Cost Control... The relevant header: H10:I11 merged... H12 contains DONE".
    # This implies for Cost Control, column H is DONE Qty?
    # But for Action 1 (Target=Cost Control), we write to Week Blocks.
    # If Week 1 is at H in Report, does it match H in Cost Control?
    # If Cost Control has "DONE" at H, then it can't have "Week 1" at H.
    
    # Maybe Cost Control has Weeks at M+? And Report has Weeks at H?
    # Or maybe "DONE" is part of the weekly block structure? No.
    # "From beginning... Q-ty" is cumulative to date.
    
    # Let's assume Cost Control layout:
    # C: Description
    # H: DONE Qty (Col 8)
    # M+: Week Blocks.
    
    # But for Action 1, we match "Week 1".
    # I will put Week 1 at M (13) in Cost Control.
    # And "DONE" at H (8).
    
    ws.merge_cells("H10:I11")
    ws["H10"] = "From beginning of construction     Q-ty"
    ws["H12"] = "DONE"
    
    # Week 1 at M (13) -> M,N,O,P,Q
    ws.merge_cells("M10:Q10")
    ws["M10"] = "Week 1 - Jan 1"
    for i, h in enumerate(headers):
        ws.cell(row=13, column=13+i, value=h)
    
    # Data
    ws["C14"] = "Section Alpha"
    ws["C15"] = "Concrete"
    ws["H15"] = 500 # DONE value
    
    ws["C16"] = "Steel"
    ws["H16"] = 1200
    
    wb.save("synthetic_cost_control.xlsx")

def create_ppc():
    wb = Workbook()
    ws = wb.active
    ws.title = "PPC Sheet 1"
    
    # Headers
    ws["B1"] = "Material Name"
    ws["G1"] = "Quantity"
    
    # Data
    ws["B2"] = "Concrete"
    ws["G2"] = 0 
    
    ws["B3"] = "Steel"
    ws["G3"] = 100 
    
    ws2 = wb.create_sheet("PPC Sheet 2")
    ws2["B1"] = "Material Name"
    ws2["G1"] = "Quantity"
    ws2["B2"] = "Concrete" 
    
    wb.save("synthetic_ppc.xlsx")

if __name__ == "__main__":
    create_report()
    create_cost_control()
    create_ppc()
    print("Files created.")
