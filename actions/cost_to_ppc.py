from typing import List, Dict, Tuple, Any
from openpyxl.worksheet.worksheet import Worksheet
from actions.base import TransferAction
from parsers import SectionParser, ParsedRow
from excel_layout import find_data_start_row

COL_DONE_IDX_SRC = 8  # Column H
COL_MAT_IDX_TGT = 2   # Column B (PPC)
COL_QTY_IDX_TGT = 7   # Column G (PPC)

class TransferDiffPPC:
    def __init__(self, key, sheet_name, src_val, tgt_val, action, reason, row_idx_tgt):
        self.key = key
        self.sheet_name = sheet_name
        self.src_val = src_val
        self.tgt_val = tgt_val
        self.action = action
        self.reason = reason
        self.row_idx_tgt = row_idx_tgt

class CostToPPCAction(TransferAction):
    def __init__(self, source_sheet: Worksheet, target_sheets: List[Worksheet], 
                 use_section_matching=False, overwrite_formulas=False):
        self.source = source_sheet
        self.target_sheets = target_sheets
        self.use_section_matching = use_section_matching
        self.overwrite_formulas = overwrite_formulas

    def _parse_source(self) -> Dict[Tuple[str, str], float]:
        """
        Parses Source (Cost Control).
        Key: (section, material). Value: DONE Qty.
        """
        # "From beginning... Q-ty" at H10:I11. DONE at H12. Data starts below.
        # Use simple SectionParser.
        # Data Col to check for emptiness: H (8).
        parser = SectionParser(self.source, key_col_idx=3, start_row=13) # Materials in C (3)
        parsed_rows = parser.parse(data_cols=[COL_DONE_IDX_SRC], value_mapping={'done': COL_DONE_IDX_SRC})
        
        data = {}
        for row in parsed_rows:
            key = row.composite_key()
            val = row.values['done']
            
            if val is not None:
                if key in data:
                    data[key] += val
                else:
                    data[key] = val
        return data

    def _parse_target_sheet(self, sheet: Worksheet) -> Dict[Any, List[int]]:
        """
        Parses a single Target PPC sheet.
        Returns Index: Key -> List[row_idx].
        Key is (section, material) if use_section_matching else material_only_normalized.
        """
        # Materials in Column B (2). Qty in G (7).
        parser = SectionParser(sheet, key_col_idx=COL_MAT_IDX_TGT, start_row=2)
        parsed_rows = parser.parse(
            data_cols=[COL_QTY_IDX_TGT], 
            value_mapping={},
            detect_sections=self.use_section_matching
        )
        
        index = {}
        for row in parsed_rows:
            if self.use_section_matching:
                key = row.composite_key()
            else:
                key = row.key # Just material name normalized
            
            if key not in index:
                index[key] = []
            index[key].append(row.row_idx)
        return index

    def analyze(self) -> List[TransferDiffPPC]:
        all_diffs = []
        src_data = self._parse_source()
        
        for tgt_sheet in self.target_sheets:
            tgt_index = self._parse_target_sheet(tgt_sheet)
            
            # Iterate source data
            for (src_sec, src_mat), src_val in src_data.items():
                
                # Determine lookup key
                if self.use_section_matching:
                    lookup_key = (src_sec, src_mat)
                else:
                    lookup_key = src_mat # Normalized in parser
                
                if lookup_key not in tgt_index:
                    continue
                    
                rows = tgt_index[lookup_key]
                # "If a material matches in multiple sheets, write to all matches"
                # Here we are inside loop for one sheet. So we write to all matches in this sheet.
                
                for r in rows:
                    curr_val = tgt_sheet.cell(row=r, column=COL_QTY_IDX_TGT).value
                    
                    action = "Write"
                    reason = "Update"
                    
                    # Check Formula
                    cell_tgt = tgt_sheet.cell(row=r, column=COL_QTY_IDX_TGT)
                    if str(cell_tgt.value).startswith('=') and not self.overwrite_formulas:
                        action = "Skip"
                        reason = "Formula protected"
                        
                    diff = TransferDiffPPC(
                        key=f"{src_sec} / {src_mat}",
                        sheet_name=tgt_sheet.title,
                        src_val=src_val,
                        tgt_val=curr_val,
                        action=action,
                        reason=reason,
                        row_idx_tgt=r
                    )
                    all_diffs.append(diff)
                    
        return all_diffs

    def execute(self, diffs: List[TransferDiffPPC]):
        count = 0
        sheet_map = {ws.title: ws for ws in self.target_sheets}
        
        for diff in diffs:
            if diff.action == "Write":
                ws = sheet_map.get(diff.sheet_name)
                if ws and diff.src_val is not None:
                    ws.cell(row=diff.row_idx_tgt, column=COL_QTY_IDX_TGT).value = diff.src_val
                    count += 1
        return count
