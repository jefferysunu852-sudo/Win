from typing import List, Dict, Tuple, Optional, Any
from openpyxl.worksheet.worksheet import Worksheet
import streamlit as st
from actions.base import TransferAction
from excel_layout import WeekBlock, find_data_start_row, KEY_COL_IDX, extract_week_number
from parsers import SectionParser, ParsedRow, normalize_string

class TransferDiff:
    """Represents a single row transfer action explanation."""
    def __init__(self, key: Tuple[str, str], week_label, src_vals, tgt_vals, action, reason, row_idx_tgt=None):
        self.key = key # (Section, Material)
        self.week_label = week_label
        self.src_planned, self.src_actual, self.src_timesheet = src_vals
        self.tgt_planned, self.tgt_actual, self.tgt_timesheet = tgt_vals
        self.action = action 
        self.reason = reason
        self.row_idx_tgt = row_idx_tgt

class ReportToCostAction(TransferAction):
    def __init__(self, source_sheet: Worksheet, target_sheet: Worksheet, 
                 week_pairs: List[Tuple[WeekBlock, WeekBlock]],
                 overwrite_formulas=False, write_first_match_only=False):
        self.source = source_sheet
        self.target = target_sheet
        self.week_pairs = week_pairs
        self.overwrite_formulas = overwrite_formulas
        self.write_first_match_only = write_first_match_only

    def _get_check_cols(self, sheet: Worksheet) -> List[int]:
        """Returns a list of columns to check for emptiness (to detect Sections).
        Heuristic: Check columns D to ~K (row 13 + ~10 cols).
        Or better: Check all columns that are part of detected week blocks?
        Simple robust approach: Check columns 4 to 15.
        """
        # User example: D13:K13 empty.
        return list(range(4, 15))

    def _aggregate_source_data(self, week_block: WeekBlock) -> Dict[Tuple[str, str], Tuple[float, float, float]]:
        """
        Parses source sheet for a specific week block.
        Returns: { (section, material): (planned, actual, timesheet) }
        Summing duplicates if they exist within the same section.
        """
        # Define value mapping for this week
        val_map = {
            'planned': week_block.col_planned_qty,
            'actual': week_block.col_actual_qty,
            'timesheet': week_block.col_timesheet
        }
        
        parser = SectionParser(self.source, start_row=find_data_start_row(self.source))
        # Check cols: use the week block columns themselves + some buffers?
        # Actually, let's use the generic check_cols for section detection to be safe.
        check_cols = self._get_check_cols(self.source)
        
        # Merge week block columns into check_cols just to be sure we don't accidentally treat a data row as a header
        # if D:K are empty but Week 50 (Col AZ) has data.
        # But `SectionParser` logic is: "If ANY of data_cols has number -> Not Header".
        # So we should pass ALL columns that might contain data.
        check_cols = list(set(check_cols + [week_block.col_planned_qty, week_block.col_actual_qty]))
        
        parsed_rows = parser.parse(data_cols=check_cols, value_mapping=val_map)
        
        data = {}
        for row in parsed_rows:
            key = row.composite_key()
            p = row.values['planned']
            a = row.values['actual']
            t = row.values['timesheet']
            
            # Helper to sum optionals
            def add_opt(v1, v2):
                if v1 is None and v2 is None: return None
                return (v1 or 0.0) + (v2 or 0.0)

            if key in data:
                ep, ea, et = data[key]
                data[key] = (add_opt(ep, p), add_opt(ea, a), add_opt(et, t))
            else:
                data[key] = (p, a, t)
                
        return data

    def _build_target_index(self) -> Dict[Tuple[str, str], List[int]]:
        """
        Scans target sheet and returns index: (section, material) -> [row_indices]
        """
        parser = SectionParser(self.target, start_row=find_data_start_row(self.target))
        # For target, what columns to check?
        check_cols = self._get_check_cols(self.target)
        
        # Parsing with empty value mapping just to get structure
        parsed_rows = parser.parse(data_cols=check_cols, value_mapping={})
        
        index = {}
        for row in parsed_rows:
            key = row.composite_key()
            if key not in index:
                index[key] = []
            index[key].append(row.row_idx)
        return index

    def analyze(self) -> List[TransferDiff]:
        all_diffs = []
        tgt_index = self._build_target_index()

        for src_wk, tgt_wk in self.week_pairs:
            src_data = self._aggregate_source_data(src_wk)
            
            for key, (sp, sa, st) in src_data.items():
                if key not in tgt_index:
                    # Optional: We could try falling back to just Material Name match if Section fails?
                    # The prompt says: "Match matching rows by (section, material) rather than just material."
                    # "If target sheet has duplicates...".
                    # Let's stick to strict section matching first.
                    
                    # FALLBACK: If NO match found by section, try matching by Material only?
                    # "Match rows by (section, material)" implies strict requirement.
                    continue
                
                target_rows = tgt_index[key]
                if self.write_first_match_only:
                    target_rows = target_rows[:1]
                
                for r in target_rows:
                    # Read current target vals
                    # Note: We read simply from cells
                    tp = self.target.cell(row=r, column=tgt_wk.col_planned_qty).value
                    ta = self.target.cell(row=r, column=tgt_wk.col_actual_qty).value
                    tt = self.target.cell(row=r, column=tgt_wk.col_timesheet).value
                    
                    # Logic
                    action = "Write"
                    reason = "Update"
                    
                    # Formula Check
                    cells = [
                        self.target.cell(row=r, column=tgt_wk.col_planned_qty),
                        self.target.cell(row=r, column=tgt_wk.col_actual_qty),
                        self.target.cell(row=r, column=tgt_wk.col_timesheet)
                    ]
                    has_formula = any(str(c.value).startswith('=') for c in cells if c.value)
                    
                    if has_formula and not self.overwrite_formulas:
                        action = "Skip"
                        reason = "Target formula protected"
                    
                    # Only write if source has data? 
                    # "Only write non-blank numeric source cells"
                    if sp is None and sa is None and st is None:
                        # Nothing to transfer
                        continue
                    
                    diff = TransferDiff(
                        key, src_wk.label,
                        (sp, sa, st), (tp, ta, tt),
                        action, reason, row_idx_tgt=r
                    )
                    all_diffs.append(diff)
        return all_diffs

    def execute(self, diffs: List[TransferDiff], tgt_week_bfs: Dict[str, WeekBlock]):
        """
        tgt_week_bfs: map of week_label -> WeekBlock (to know columns)
        """
        count = 0
        for diff in diffs:
            if diff.action == "Write" and diff.row_idx_tgt:
                tgt_wk = tgt_week_bfs.get(diff.week_label)
                if not tgt_wk:
                    continue
                
                r = diff.row_idx_tgt
                sp, sa, st = diff.src_planned, diff.src_actual, diff.src_timesheet
                
                if sp is not None:
                    self.target.cell(row=r, column=tgt_wk.col_planned_qty).value = sp
                if sa is not None:
                    self.target.cell(row=r, column=tgt_wk.col_actual_qty).value = sa
                if st is not None:
                    self.target.cell(row=r, column=tgt_wk.col_timesheet).value = st
                count += 1
        return count
