from openpyxl import Workbook

from excel_layout import (
    detect_week_blocks,
    normalize_header,
    parse_number,
)


def test_normalize_header():
    assert normalize_header("Man/hour") == "man hour"
    assert normalize_header("Timesheet H") == "timesheet h"
    assert normalize_header("q-ty") == "q ty"


def test_parse_number():
    assert parse_number("1,5") == 1.5
    assert parse_number("1 234,5") == 1234.5
    assert parse_number(2) == 2.0
    assert parse_number("") is None


def test_detect_week_blocks_with_merged_header():
    wb = Workbook()
    ws = wb.active
    ws.merge_cells(start_row=10, start_column=5, end_row=10, end_column=9)
    ws.cell(row=10, column=5, value="WK8")
    headers = ["q-ty", "Man/hour", "q-ty", "Man/hour", "Timesheet"]
    for idx, header in enumerate(headers, start=5):
        ws.cell(row=13, column=idx, value=header)

    blocks = detect_week_blocks(ws)
    assert len(blocks) == 1
    assert blocks[0].label == "WK8"
    assert blocks[0].start_col == 5
    assert blocks[0].end_col == 9
