import pytest
import sys
import os
from openpyxl import Workbook

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from transfer import TransferManager, TransferDiff, WeekBlock
from excel_layout import KEY_COL_IDX

@pytest.fixture
def mock_sheets():
    wb = Workbook()
    ws_src = wb.active
    ws_src.title = "Source"
    ws_tgt = wb.create_sheet("Target")
    return ws_src, ws_tgt

def test_transfer_robust_multi_row(mock_sheets):
    src, tgt = mock_sheets
    wb_src = WeekBlock("WK1", 5, 9, "WK1")
    wb_tgt = WeekBlock("WK1", 5, 9, "WK1")
    
    # Source Data: 3 items
    keys = ["Item A", "Item B", "Item C"]
    for i, k in enumerate(keys):
        r = 15 + i
        src.cell(row=r, column=KEY_COL_IDX).value = k
        src.cell(row=r, column=wb_src.col_planned_qty).value = (i+1)*10
        src.cell(row=r, column=wb_src.col_actual_qty).value = (i+1)*5
    
    # Target Data: 3 items mixed order
    tgt_keys = ["Item C", "Item A", "Item B"]
    for i, k in enumerate(tgt_keys):
        r = 20 + i
        tgt.cell(row=r, column=KEY_COL_IDX).value = k
        
    mgr = TransferManager(src, tgt)
    
    # Test diff generation
    diffs = mgr.transfer_weeks([(wb_src, wb_tgt)])
    
    assert len(diffs) == 3
    # Check Item A (Source row 15, Val 10 -> Target row 21)
    diff_a = next(d for d in diffs if d.key == "item a")
    assert diff_a.src_planned == 10.0
    assert diff_a.row_idx_tgt == 21
    
    # Check Item C (Source row 17, Val 30 -> Target row 20)
    diff_c = next(d for d in diffs if d.key == "item c")
    assert diff_c.src_planned == 30.0
    assert diff_c.row_idx_tgt == 20

def test_duplicate_target_write_all(mock_sheets):
    src, tgt = mock_sheets
    wb_src = WeekBlock("WK1", 5, 9, "WK1")
    wb_tgt = WeekBlock("WK1", 5, 9, "WK1")
    
    # Source: Item A = 100
    src.cell(row=15, column=KEY_COL_IDX).value = "Item A"
    src.cell(row=15, column=wb_src.col_planned_qty).value = 100
    
    # Target: Item A appears TWICE
    tgt.cell(row=20, column=KEY_COL_IDX).value = "Item A"
    tgt.cell(row=30, column=KEY_COL_IDX).value = "Item A"
    
    # Method 1: Write All (Default behavior in our revised code logic if toggle is False)
    # The toggle in UI is 'Write First Only', defaulting to False. 
    # The Class init arg is `write_first_match_only=False` by default.
    mgr = TransferManager(src, tgt, write_first_match_only=False)
    diffs = mgr.transfer_weeks([(wb_src, wb_tgt)])
    
    assert len(diffs) == 2
    assert {d.row_idx_tgt for d in diffs} == {20, 30}
    
    # Method 2: Write First Only
    mgr_first = TransferManager(src, tgt, write_first_match_only=True)
    diffs_first = mgr_first.transfer_weeks([(wb_src, wb_tgt)])
    
    assert len(diffs_first) == 1
    assert diffs_first[0].row_idx_tgt == 20 # First one

def test_source_aggregation_blanks(mock_sheets):
    src, tgt = mock_sheets
    wb_src = WeekBlock("WK1", 5, 9, "WK1")
    
    # Row 15: Item A, Plan=10
    src.cell(row=15, column=KEY_COL_IDX).value = "Item A"
    src.cell(row=15, column=wb_src.col_planned_qty).value = 10
    
    # Row 16: Item A, Plan=None (Blank)
    src.cell(row=16, column=KEY_COL_IDX).value = "Item A"
    # Leave blank
    
    mgr = TransferManager(src, tgt)
    data = mgr.build_source_aggregate(wb_src)
    
    # Should be 10 + 0 = 10
    assert data["item a"][0] == 10.0
    
    # Row 17: Item B, Plan=None
    src.cell(row=17, column=KEY_COL_IDX).value = "Item B"
    
    data = mgr.build_source_aggregate(wb_src)
    assert data["item b"][0] is None # Should stay None if all blank
