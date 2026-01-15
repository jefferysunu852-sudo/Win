import pytest
from openpyxl import load_workbook
from actions.report_to_cost import ReportToCostAction
from actions.cost_to_ppc import CostToPPCAction
from excel_layout import WeekBlock

def test_report_to_cost_action():
    # Load synthetic
    wb_src = load_workbook("synthetic_report.xlsx", data_only=False)
    wb_tgt = load_workbook("synthetic_cost_control.xlsx", data_only=False)
    
    ws_src = wb_src["Report"]
    ws_tgt = wb_tgt["Cost Control"]
    
    # Mock Week Blocks
    # Synthetic Report: Week 1 at H (8). Width 5 -> 8,9,10,11,12.
    src_wb = WeekBlock("Week 1 - Jan 1", 8, 12, "Week 1 - Jan 1")
    
    # Synthetic Cost Control: Week 1 at M (13). Width 5 -> 13,14,15,16,17.
    tgt_wb = WeekBlock("Week 1 - Jan 1", 13, 17, "Week 1 - Jan 1")
    
    pairs = [(src_wb, tgt_wb)]
    
    action = ReportToCostAction(ws_src, ws_tgt, pairs)
    diffs = action.analyze()
    
    # In synthetic_report:
    # Section Alpha / Steel -> Plan 100, Act 50, Time 10
    # In synthetic_cost_control:
    # Section Alpha / Steel -> Exists but empty
    
    # Find diff for Steel
    diff = next((d for d in diffs if d.key[1] == "steel"), None)
    assert diff is not None
    assert diff.key[0] == "Section Alpha"
    assert diff.src_planned == 100.0
    assert diff.src_actual == 50.0
    assert diff.src_timesheet == 10.0
    assert diff.action == "Write"

def test_cost_to_ppc_action():
    wb_src = load_workbook("synthetic_cost_control.xlsx", data_only=True) # Data only for calculation
    wb_tgt = load_workbook("synthetic_ppc.xlsx", data_only=False)
    
    ws_src = wb_src["Cost Control"]
    tgt_sheets = [wb_tgt["PPC Sheet 1"], wb_tgt["PPC Sheet 2"]]
    
    action = CostToPPCAction(ws_src, tgt_sheets, use_section_matching=False)
    diffs = action.analyze()
    
    # Expect:
    # Concrete -> 500 (Source H15)
    # Matches PPC Sheet 1 (B2) and Sheet 2 (B2)
    
    concrete_diffs = [d for d in diffs if "concrete" in d.key]
    assert len(concrete_diffs) == 2
    for d in concrete_diffs:
        assert d.src_val == 500.0
        assert d.action == "Write"
        
    # Steel -> 1200 (Source H16)
    # Matches PPC Sheet 1 (B3)
    steel_diffs = [d for d in diffs if "steel" in d.key]
    assert len(steel_diffs) == 1
    assert steel_diffs[0].src_val == 1200.0
    assert steel_diffs[0].sheet_name == "PPC Sheet 1"
