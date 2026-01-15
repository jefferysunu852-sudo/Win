import pytest
from openpyxl import Workbook
from parsers import SectionParser, parse_number, normalize_string

def test_parse_number():
    assert parse_number("1,5") == 1.5
    assert parse_number("1 234,56") == 1234.56
    assert parse_number("24") == 24.0
    assert parse_number(None) is None
    assert parse_number("") is None

def test_section_parser():
    wb = Workbook()
    ws = wb.active
    
    # Setup data
    # Row 1: Section Header 1 (Text in C, Empty D)
    ws["C1"] = "Section Alpha"
    # Row 2: Data (Text in C, Number in D)
    ws["C2"] = "Material 1"
    ws["D2"] = 100
    # Row 3: Data
    ws["C3"] = "Material 2"
    ws["D3"] = 200
    
    # Row 4: Section Header 2
    ws["C4"] = "Section Beta" 
    # D4 is None
    
    # Row 5: Data
    ws["C5"] = "Material 3"
    ws["D5"] = 300
    
    parser = SectionParser(ws, key_col_idx=3, start_row=1)
    # Check emptiness in col 4 (D)
    # Value mapping: Alias 'val' -> Col 4
    rows = parser.parse(data_cols=[4], value_mapping={'val': 4})
    
    assert len(rows) == 3
    
    # Check Row 1 -> Section Alpha (no record yielded because it's a header)
    
    # Check Header capture
    assert rows[0].section == "Section Alpha"
    assert rows[0].key == "material 1"
    assert rows[0].values['val'] == 100.0
    
    assert rows[1].section == "Section Alpha"
    assert rows[1].key == "material 2"
    
    assert rows[2].section == "Section Beta"
    assert rows[2].key == "material 3"
