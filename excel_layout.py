import re
from typing import List, Optional
from openpyxl.worksheet.worksheet import Worksheet
from parsers import normalize_string

# Constants config
HEADER_ROW_YEAR = 8
HEADER_ROW_WEEK = 10
HEADER_ROW_SUB = 13
KEY_COL_IDX = 3  # Column C is index 3 (1-based)

# Expected patterns
EXPECTED_SUBHEADERS = [
    r"q[\-\'\s]?ty",           # 1. Planned Qty
    r"(m|man)[\-\s\/\\]?hour", # 2. Planned MH
    r"q[\-\'\s]?ty",           # 3. Actual Qty
    r"(m|man)[\-\s\/\\]?hour", # 4. Actual MH
    r"time\s?sheet"            # 5. Timesheet
]

class WeekBlock:
    """Represents a detected week block in the Excel sheet."""
    def __init__(self, label: str, start_col: int, end_col: int, raw_label: str):
        self.label = label
        self.start_col = start_col  # 1-based index
        self.end_col = end_col      # 1-based index
        self.raw_label = raw_label

    def __repr__(self):
        return f"WeekBlock(label='{self.label}', cols={self.start_col}-{self.end_col})"

    @property
    def col_planned_qty(self):
        return self.start_col + 0

    @property
    def col_actual_qty(self):
        return self.start_col + 2

    @property
    def col_timesheet(self):
        return self.start_col + 4

def normalize_header(s: str) -> str:
    """Slightly more aggressive normalization for sub-headers."""
    if not s:
        return ""
    return str(s).lower().strip()

def check_subheader_pattern(sheet: Worksheet, start_col_idx: int) -> bool:
    """
    Checks if columns start_col_idx to start_col_idx+4 match the expected pattern.
    """
    for i, pattern in enumerate(EXPECTED_SUBHEADERS):
        col_idx = start_col_idx + i
        cell = sheet.cell(row=HEADER_ROW_SUB, column=col_idx)
        val = normalize_header(cell.value)
        # Check regex match
        if not re.search(pattern, val):
            return False
    return True

def extract_week_number(label: str) -> Optional[int]:
    """Extracts the first number found in a week label as a fallback."""
    match = re.search(r'\d+', label)
    if match:
        return int(match.group(0))
    return None

def detect_week_blocks(sheet: Worksheet) -> List[WeekBlock]:
    """
    Scans row 10 (merged cells) to find valid week blocks.
    Validation:
    1. Must be merged across exactly 5 columns.
    2. Sub-headers in row 13 must match EXPECTED_SUBHEADERS.
    """
    blocks = []
    
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row <= HEADER_ROW_WEEK <= merged_range.max_row:
            start_col = merged_range.min_col
            end_col = merged_range.max_col
            width = end_col - start_col + 1
            
            if width != 5:
                continue
                
            if check_subheader_pattern(sheet, start_col):
                top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                raw_label = str(top_left_cell.value) if top_left_cell.value else ""
                
                norm_label = normalize_string(raw_label)
                if not norm_label:
                    continue
                    
                blocks.append(WeekBlock(norm_label, start_col, end_col, raw_label))
    
    blocks.sort(key=lambda x: x.start_col)
    
    if not blocks:
        # Fallback: Scan row 13 for patterns directly
        max_col = sheet.max_column
        col = 1
        while col <= max_col - 4:
            if check_subheader_pattern(sheet, col):
                val = sheet.cell(row=HEADER_ROW_WEEK, column=col).value
                raw_label = str(val) if val else f"Week Col {col}"
                blocks.append(WeekBlock(normalize_string(raw_label), col, col+4, raw_label))
                col += 5
            else:
                col += 1
                
    return blocks

def find_data_start_row(sheet: Worksheet) -> int:
    """
    Finds first row after header that has a value in Key Column (C).
    Default search starts from row 14.
    """
    start_search = HEADER_ROW_SUB + 1
    max_search = min(start_search + 50, sheet.max_row + 1)
    
    for r in range(start_search, max_search):
        val = sheet.cell(row=r, column=KEY_COL_IDX).value
        if val:
            norm_val = normalize_string(str(val))
            if "description" not in norm_val:
                return r
    return start_search
