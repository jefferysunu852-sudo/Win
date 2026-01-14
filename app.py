"""Streamlit app for transferring weekly data between REPORT and COST CONTROL workbooks.

Usage:
1) Upload the REPORT (source) and COST CONTROL (target) Excel files.
2) Click Analyze to detect week blocks and preview matching details.
3) Select a week label and click Transfer week to update COST CONTROL.
4) Download the updated workbook as cost_control_updated.xlsx.

The app preserves formulas and formatting using openpyxl and provides detailed
logging for traceability.
"""

from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from excel_layout import (
    WeekBlock,
    build_week_map,
    detect_week_blocks,
    find_data_start_row,
    match_week_label,
)
from transfer import (
    TransferResult,
    TransferSettings,
    analyze_transfer,
    transfer_week,
)


st.set_page_config(page_title="Weekly Transfer", layout="wide")

st.title("Weekly data transfer")
st.caption("Upload REPORT and COST CONTROL workbooks to transfer weekly data. (LT: kelti savaitinius duomenis)")

st.sidebar.header("Inputs")

source_file = st.sidebar.file_uploader("REPORT workbook", type=["xlsx"], key="source")
target_file = st.sidebar.file_uploader("COST CONTROL workbook", type=["xlsx"], key="target")

if source_file:
    st.session_state["source_bytes"] = source_file.getvalue()
if target_file:
    st.session_state["target_bytes"] = target_file.getvalue()

source_bytes = st.session_state.get("source_bytes")
target_bytes = st.session_state.get("target_bytes")

source_wb = load_workbook(BytesIO(source_bytes)) if source_bytes else None
target_wb = load_workbook(BytesIO(target_bytes)) if target_bytes else None

source_sheet_name = None
target_sheet_name = None

if source_wb:
    source_sheet_name = st.sidebar.selectbox("Source sheet", source_wb.sheetnames)
if target_wb:
    target_sheet_name = st.sidebar.selectbox("Target sheet", target_wb.sheetnames)

write_all_duplicates = st.sidebar.checkbox("Write to all duplicate target keys", value=False)
overwrite_formulas = st.sidebar.checkbox("Overwrite formulas", value=False)

analyze_clicked = st.sidebar.button("Analyze")
transfer_clicked = st.sidebar.button("Transfer week")

week_label = None
week_options: List[str] = []
source_blocks: List[WeekBlock] = []
target_blocks: List[WeekBlock] = []
source_block_map: Dict[str, WeekBlock] = {}
target_block_map: Dict[str, WeekBlock] = {}

if source_wb and source_sheet_name:
    source_ws = source_wb[source_sheet_name]
    source_blocks = detect_week_blocks(source_ws)
    source_block_map = build_week_map(source_blocks)
    week_options = [block.label for block in source_blocks]

if target_wb and target_sheet_name:
    target_ws = target_wb[target_sheet_name]
    target_blocks = detect_week_blocks(target_ws)
    target_block_map = build_week_map(target_blocks)

if week_options:
    week_label = st.sidebar.selectbox("Week label", week_options)

settings = TransferSettings(
    write_all_duplicates=write_all_duplicates,
    overwrite_formulas=overwrite_formulas,
)

if analyze_clicked and source_wb and target_wb and week_label:
    source_ws = source_wb[source_sheet_name]
    target_ws = target_wb[target_sheet_name]
    source_block = match_week_label(week_label, source_block_map)
    target_block = match_week_label(week_label, target_block_map)

    if not source_block:
        st.error("Selected week not found in REPORT.")
    elif not target_block:
        st.error("Selected week not found in COST CONTROL.")
    else:
        analysis = analyze_transfer(source_ws, target_ws, source_block, target_block, settings)
        st.subheader("Detected weeks")
        st.write("**REPORT**")
        st.table(
            [
                {
                    "Week": block.label,
                    "Columns": f"{get_column_letter(block.start_col)}-{get_column_letter(block.end_col)}",
                }
                for block in source_blocks
            ]
        )
        st.write("**COST CONTROL**")
        st.table(
            [
                {
                    "Week": block.label,
                    "Columns": f"{get_column_letter(block.start_col)}-{get_column_letter(block.end_col)}",
                }
                for block in target_blocks
            ]
        )

        st.subheader("Preview summary")
        st.write(
            {
                "Matched keys": analysis.matched_keys,
                "Missing target keys": analysis.missing_target_keys,
                "Duplicate source keys": analysis.duplicate_source_keys,
                "Duplicate target keys": analysis.duplicate_target_keys,
                "Skipped formula cells": analysis.skipped_formula_cells,
            }
        )
        st.subheader("Diff preview")
        st.dataframe(pd.DataFrame(analysis.diff_rows))

if transfer_clicked and source_wb and target_wb and week_label:
    source_ws = source_wb[source_sheet_name]
    target_ws = target_wb[target_sheet_name]
    source_block = match_week_label(week_label, source_block_map)
    target_block = match_week_label(week_label, target_block_map)

    if not source_block:
        st.error("Selected week not found in REPORT.")
    elif not target_block:
        st.error("Selected week not found in COST CONTROL.")
    else:
        result: TransferResult = transfer_week(source_ws, target_ws, source_block, target_block, settings)
        st.success("Transfer completed.")
        st.subheader("Transfer summary")
        st.write(
            {
                "Written cells": result.written_cells,
                "Matched keys": result.matched_keys,
                "Missing target keys": result.missing_target_keys,
                "Duplicate source keys": result.duplicate_source_keys,
                "Duplicate target keys": result.duplicate_target_keys,
                "Skipped formula cells": result.skipped_formula_cells,
            }
        )
        st.subheader("Diff results")
        st.dataframe(pd.DataFrame(result.diff_rows))
        st.subheader("Log")
        st.text("\n".join(result.logs))

        output = BytesIO()
        target_wb.save(output)
        output.seek(0)
        st.download_button(
            "Download updated COST CONTROL",
            data=output,
            file_name="cost_control_updated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

if not source_bytes or not target_bytes:
    st.info("Upload both workbooks to begin. (LT: kelti abu failus)")
